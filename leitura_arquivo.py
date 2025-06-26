import pandas as pd
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from fuzzywuzzy import fuzz
import io

# Realizar a busca de informaçao sobre se a nota é ST nas Operacoes Internas
with open("arquivo.json", "r", encoding="utf-8") as f:
    dados_caderno = json.load(f)  # Agora está correto


def ler_planilha(planilha):
    # df = pd.read_excel('input/planilha_input_terra_util.xlsx')
    df = pd.read_excel(planilha)

    # Primeiro criar as colunas para analise
    colunas_criar = ["TEM CONVÊNIO OU PROTOCOLO DE ICMS?", "SUBSTITUTO TRIBUTÁRIO OPERAÇÕES INTERNAS?",
                    "O CÁLCULO ESTÁ CORRETO?", "MVA DA NF-e", "MVA DA LEGISLAÇÃO", "CFOP DA NF-e",
                    "CFOP DA LEGISLAÇÃO", "Análise da Metrópole", "RECOMENDAÇÕES", "NF-e COMPLEMENTAR",
                    "BC ICMS RET METRÓPOLE", "ICMS ST METRÓPOLE", "DIFERENÇA BC METRÓPOLE", 
                    "DIFERENÇA ICMS ST METRÓPOLE"]

    # Adicionando as novas colunas com valores padrão (ex.: None)
    for col in colunas_criar:
        if col in ["BC ICMS RET METRÓPOLE", "ICMS ST METRÓPOLE", "DIFERENÇA BC METRÓPOLE", 
                    "DIFERENÇA ICMS ST METRÓPOLE"]:
            df[col] = 0.0
        else:
            df[col] = ""

    # Realizar o calculo da base de ST

    df.loc[df[" Cod. Fiscal"].isin([5403, 6404]), "BC ICMS RET METRÓPOLE"] = round((df["Valor Frete"] + df["Vlr Seguro"] + df["Vlr Despesas"] - df["Vlr Desconto"] + df["Vlr Total"]) * (1+df["Margem "]/100), 2)
    df.loc[df[" Cod. Fiscal"].isin([5403, 6404]), "ICMS ST METRÓPOLE"] = round((df["BC ICMS RET METRÓPOLE"]*0.2)-df["Valor ICMS"], 2)

    df.loc[df[" Cod. Fiscal"].isin([5403, 6404]), "DIFERENÇA BC METRÓPOLE"] = df["Vlr Base Ret"] - df["BC ICMS RET METRÓPOLE"]
    df.loc[df[" Cod. Fiscal"].isin([5403, 6404]), "DIFERENÇA ICMS ST METRÓPOLE"] = df["ICMS ST METRÓPOLE"] - df["Vlr ICMS Ret"]

    # zerar os valores na coluna de diferenças quando -0.1 <= x <= 0.1
    df.loc[(df["DIFERENÇA ICMS ST METRÓPOLE"] >= -0.1) & (df["DIFERENÇA ICMS ST METRÓPOLE"] <= 0.1), "DIFERENÇA ICMS ST METRÓPOLE"] = 0.0
    df.loc[(df["DIFERENÇA BC METRÓPOLE"] >= -0.1) & (df["DIFERENÇA BC METRÓPOLE"] <= 0.1), "DIFERENÇA BC METRÓPOLE"] = 0.0
    
    # VERIFICAR SE O PRODUTO TEM CONVENIO OU PROTOCOLO DE ICMS PARA CFOP 6404
    
    df["TEM CONVÊNIO OU PROTOCOLO DE ICMS?"] = df.apply(verificar_convenio, axis=1)

    df["SUBSTITUTO TRIBUTÁRIO OPERAÇÕES INTERNAS?"] = df.apply(verificar_produto_caderno, axis=1)

    df["MVA DA NF-e"] = df["Margem "]
    df["MVA DA NF-e"] = df["MVA DA NF-e"].map(lambda x: f"{x:.2f}".replace('.', ',') + "%")

    df["MVA DA LEGISLAÇÃO"] = df.apply(mva_legislacao, axis=1)
    df["MVA DA LEGISLAÇÃO"] = df["MVA DA LEGISLAÇÃO"].map(lambda x: f"{x:.2f}".replace('.', ',') + "%" if str(x) != 'Indicar MVA' else x)
    
    df["CFOP DA NF-e"] = df[" Cod. Fiscal"]

    df["CFOP DA LEGISLAÇÃO"] = df.apply(cfop_legislacao, axis=1)
    
    df["O CÁLCULO ESTÁ CORRETO?"] = df.apply(verificar_calculo_correto, axis=1)

    df["RECOMENDAÇÕES"] = df.apply(recomendacoes, axis=1)
    
    df["Análise da Metrópole"] = df.apply(analise_metropoles, axis=1)


    return df


def verificar_convenio(row):
    if row[" Cod. Fiscal"] == 6404:
        convenio = 'Não'
        if str(row["CEST SFT"]).strip() != '':
            for item in dados_caderno:
                if item['CEST'] is None:
                    if str(item['NCMSH']).replace('.', '').strip() in str(row["   NCM Cadastro de Produto"]):
                        if row['Estado Ref '] in item['UFDEORIGEM']:
                            convenio = 'Sim'
                            break
                elif str(item['CEST']).replace('.', '') in str(row["CEST SFT"]) or str(item['NCMSH']).replace('.', '').strip() in str(row["   NCM Cadastro de Produto"]):
                    if row['Estado Ref '] in item['UFDEORIGEM']:
                        convenio = 'Sim'
                        break
        else:
            for item in dados_caderno:
                if str(item['NCMSH']).replace('.', '').strip() in str(row["   NCM Cadastro de Produto"]):
                    if row['Estado Ref '] in item['UFDEORIGEM']:
                        convenio = 'Sim'
                        break
        # return "SIM" if (row['CEST'], row['NCM']) in dicionario_validacao else "NÃO"
        return convenio
    return "Não Aplicável"

def verificar_produto_caderno(row):
    existe_caderno = 'Não'
    if row[" Cod. Fiscal"] in [5922, 5914, 5910, 5119, 5117, 2202, 1202]:
        return 'Não'
    
    if str(row["CEST SFT"]).strip() != '':
        achou_cest = False
        for item in dados_caderno:
            if item['CEST'] is None:
                continue
            if str(item['CEST']).replace('.', '') == row["CEST SFT"].strip():
                achou_cest = True
                existe_caderno = 'Sim'
                break
        if not achou_cest:
            for item in dados_caderno:
                if str(row["   NCM Cadastro de Produto"]).replace('.', '').strip()[:4] in str(item['NCMSH']).replace('.', '').strip():
                    existe_caderno = 'Sim'
                    break
    
    else:
        for item in dados_caderno:
            if str(row["   NCM Cadastro de Produto"]).replace('.', '').strip()[:4] in str(item['NCMSH']).replace('.', '').strip():
                existe_caderno = 'Sim'
                break

    return existe_caderno

def verificar_calculo_correto(row):
    if row[" Cod. Fiscal"] == 5403 or row[" Cod. Fiscal"] == 6404:
        # if row['SUBSTITUTO TRIBUTÁRIO OPERAÇÕES INTERNAS?'] == 'Sim':
        if -0.1 <= row["DIFERENÇA ICMS ST METRÓPOLE"] <= 0.1 and row['MVA DA NF-e'] == row['MVA DA LEGISLAÇÃO']:
            return 'Sim'
        
        elif float(row['Margem ']) == 0:
            return 'Não houve cálculo'
        return 'Não'

    else:
        return 'Não houve cálculo'

def mva_legislacao(row):
    if row['SUBSTITUTO TRIBUTÁRIO OPERAÇÕES INTERNAS?'] == 'Sim':
        if row[" Cod. Fiscal"] == 6102:
            return 'Indicar MVA'
        if str(row["CEST SFT"]).strip() == '':
            achou_ncm = False
            for item in dados_caderno:
                # if item['CEST'] is None:
                    # if str(item['NCMSH']).replace('.', '').strip() in str(row["   NCM Cadastro de Produto"]):
                    if str(row["   NCM Cadastro de Produto"]).replace('.', '').strip()[:4] in str(item['NCMSH']).replace('.', '').strip():
                        achou_ncm = True
                        return item["MVAST_Interna_Atacadistas"]
            
            if not achou_ncm:
                return 0.00
                
        else:
            achou_cest = False
            for item in dados_caderno:
                if item['CEST'] is None:
                    continue
                if str(item['CEST']).replace('.', '') == row["CEST SFT"].strip():
                    achou_cest = True
                    return item["MVAST_Interna_Atacadistas"]
            if not achou_cest:
                achou_ncm = False
                for item in dados_caderno:
                    # if item['CEST'] is None:
                        # if str(item['NCMSH']).replace('.', '').strip() in str(row["   NCM Cadastro de Produto"]):
                        if str(row["   NCM Cadastro de Produto"]).replace('.', '').strip()[:4] in str(item['NCMSH']).replace('.', '').strip():
                            achou_ncm = True
                            return item["MVAST_Interna_Atacadistas"]
                
                if not achou_ncm:
                    return 0.00

    return 0.00

def cfop_legislacao(row):
    if row['CFOP DA NF-e'] == 6404:
        if row['TEM CONVÊNIO OU PROTOCOLO DE ICMS?'] == 'Não':
            return 6102
        else:
            return 6404
    
    else:
        if row['SUBSTITUTO TRIBUTÁRIO OPERAÇÕES INTERNAS?'] == 'Sim':
            if row['CFOP DA NF-e'] == 6102:
                return 6404
            elif row['CFOP DA NF-e'] == 5102:
                return 5403
            elif row['CFOP DA NF-e'] in [2411, 1411]:
                return 6404
            elif row['Estado Ref '] == 'DF':
                return row['CFOP DA NF-e']
            else:
                return row['CFOP DA NF-e']
        else:
            if row['CFOP DA NF-e'] == 5403:
                return 5102
            else:
                return row['CFOP DA NF-e']

def analise_metropoles(row):
    if row['CFOP DA NF-e'] == 5403:
        if row['SUBSTITUTO TRIBUTÁRIO OPERAÇÕES INTERNAS?'] == 'Sim':
            if row["MVA DA NF-e"] != row["MVA DA LEGISLAÇÃO"]:
                return 'Analisar'
            elif str(row["CEST"]).strip() != '' and str(row["   NCM Cadastro de Produto"]).strip()  != '':
                if -0.1 <= row['DIFERENÇA ICMS ST METRÓPOLE'] <= 0.1:
                    return 'Validado'
                else:
                    return 'Validado Parcialmente'
            else:
                if -0.1 <= row['DIFERENÇA ICMS ST METRÓPOLE'] <= 0.1:
                    return 'Validado Parcialmente'
                else:
                    return 'Validado Parcialmente'
        
        else:
            return 'Analisar'
    
    elif row['CFOP DA NF-e'] == 6404:
        if row['SUBSTITUTO TRIBUTÁRIO OPERAÇÕES INTERNAS?'] == 'Sim':
            if str(row["CEST"]).strip() != '' and str(row["   NCM Cadastro de Produto"]).strip() != '':
                if -0.1 <= row['DIFERENÇA ICMS ST METRÓPOLE'] <= 0.1:
                    if row['TEM CONVÊNIO OU PROTOCOLO DE ICMS?'] == 'Sim':
                        return 'Validado'
                    else:
                        return 'Validado Parcialmente'
                else:
                    return 'Validado Parcialmente'
            
            else:
                if -0.1 <= row['DIFERENÇA ICMS ST METRÓPOLE'] <= 0.1:
                    return 'Validado Parcialmente'
                else:
                    return 'Validado Parcialmente'
        else:
            return 'Analisar'
    elif row['CFOP DA NF-e'] in [2411, 1411]:
        return 'Analisar'
    
    else:
        if row['CFOP DA NF-e'] != row["CFOP DA LEGISLAÇÃO"]:
            if row['CFOP DA NF-e'] == 5102:
                return 'Não Validado em virtude desse produto está sujeito ao ICMS ST'
            return 'Analisar'
        
        elif row['MVA DA NF-e'] != row['MVA DA LEGISLAÇÃO']:
            return 'Analisar'
        
        elif row['RECOMENDAÇÕES'] == 'Indicar CEST':
            return 'Validado Parcialmente'
        
        else:
            return 'Validado'
        

def recomendacoes(row):
    # if row['Análise da Metrópole'] == 'Validado Parcialmente':
    if row['CFOP DA NF-e'] == 5102 and row['SUBSTITUTO TRIBUTÁRIO OPERAÇÕES INTERNAS?'] == 'Sim':
        return 'Ajustar o cadastro do produto para que nas  próximas vendas calcule  o ICMS ST, exceto se destinado a uso e consumo ou ativo imobilizado do adquirente devidamente comprovado por declaração expressa.'
    
    if str(row["CEST SFT"]).strip() == '':
        return 'Indicar CEST'
        
def teste_caderno():
    df = pd.read_excel('input/(250107) TABELA SFT ICMS ST SAIDA DF 12-2024 - TERRA ÚTIL V1.3.xlsx', sheet_name='Caderno')
    import re
    from unidecode import unidecode

    df.columns = [re.sub(r'[^a-zA-Z0-9_]', '', unidecode(col)) for col in df.columns]
    # Aplicando unidecode para remover acentos de todas as células
    df = df.applymap(lambda x: unidecode(x) if isinstance(x, str) else x)
    
    # Salvar em um arquivo JSON
    df.to_json('arquivo.json', orient='records', indent=4)

def personalizar_planilha(df):
    # wb = openpyxl.load_workbook(planilha)
    # ws = wb.active
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Dados')
    
        # Acessar a planilha ativa
        wb = writer.book
        ws = writer.sheets["Dados"]

        # Definir a cor de fundo
        cor_fundo_azul_claro = PatternFill(start_color="00F3FF", end_color="00F3FF", fill_type="solid")

        # Aplicar os estilos nas colunas que vem padrao do sistema
        for col in range(1, 50):
            ws.cell(row=1, column=col).fill = cor_fundo_azul_claro
            ws.cell(row=1, column=col).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=1, column=col).font = Font(size=8, bold=True)

            # Alterar a largura da coluna
            letra_coluna = get_column_letter(col)
            ws.column_dimensions[letra_coluna].width = 10

        ws.row_dimensions[1].height = 50

        # Definir a cor de fundo
        cor_fundo_preto = PatternFill(start_color="000000", end_color="000000", fill_type="solid")

        # Aplicar os estilos nas colunas que foram criados para analise da metropoles
        for col2 in range(50, ws.max_column + 1):
            ws.cell(row=1, column=col2).fill = cor_fundo_preto
            ws.cell(row=1, column=col2).font = Font(size=8, bold=True, color="ffffff")
            ws.cell(row=1, column=col2).alignment = Alignment(horizontal="center", vertical="center")
            
            # Alterar a largura da coluna
            letra_coluna = get_column_letter(col2)
            ws.column_dimensions[letra_coluna].width = 25

    # wb.save('teste2.xlsx')
    output.seek(0)
    return output
            
if __name__ == '__main__':
    ler_planilha('/Users/matheusbarbosa/Downloads/(250408) SFT TERRA UTIL ORIGINAL V1.0.xlsx')
    # personalizar_planilha('teste.xlsx')
    # teste_caderno()

    # descricao_tabela = "Parafusos, pinos ou pernos, roscados, porcas, tira-fundos, ganchos roscados, rebites, chavetas, cavilhas, contrapinos, arruelas (incluídas as de pressão) e artefatos semelhantes, de ferro fundido, ferro ou aço"
    # descricao_planilha = "PORCA SX 10MM MA-1,50 G2 PL                       "

    # for palavras in descricao_tabela.split(','):
    #     print(palavras)
    #     similaridade = fuzz.ratio(palavras, descricao_planilha.strip().split(' ')[0])
    #     similaridade1 = fuzz.partial_ratio(palavras, descricao_planilha.strip().split(' ')[0])
    #     similaridade2 = fuzz.token_sort_ratio(palavras, descricao_planilha.strip().split(' ')[0])
        
    #     print(similaridade, similaridade1, similaridade2)
        # print()